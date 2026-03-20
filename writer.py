import os
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import FIELDS, INPUT_COLUMNS, SOURCE_WORKBOOK, OUTPUT_WORKBOOK
from geo import GeoContext

# --- Style constants ---
FILL_HEADER     = PatternFill("solid", fgColor="404040")  # dark gray
FILL_META       = PatternFill("solid", fgColor="D9D9D9")  # light gray
FILL_PASS       = PatternFill("solid", fgColor="00B050")
FILL_MARGINAL   = PatternFill("solid", fgColor="FFC000")
FILL_FAIL       = PatternFill("solid", fgColor="FF0000")

FONT_HEADER_STD   = Font(bold=True, color="FFFFFF")
FONT_HEADER_INPUT = Font(bold=True, color="0000FF")
FONT_META         = Font(bold=False, color="404040", italic=True)


def write_to_excel(
    properties: list[dict],
    geo_context: GeoContext,
    output_path: str | None = None,
) -> str:
    if output_path is None:
        output_path = OUTPUT_WORKBOOK

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

    wb = _load_or_create_workbook()
    ws = _get_or_create_sheet(wb, "Raw_Data")

    _write_metadata_row(ws, geo_context, len(properties))
    _write_header_row(ws)
    _write_data_rows(ws, properties)
    _apply_column_widths(ws)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(FIELDS))}2"

    wb.save(output_path)
    return output_path


def _load_or_create_workbook() -> openpyxl.Workbook:
    if os.path.exists(SOURCE_WORKBOOK):
        return openpyxl.load_workbook(SOURCE_WORKBOOK)
    return openpyxl.Workbook()


def _get_or_create_sheet(wb: openpyxl.Workbook, name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)  # clear existing data, keep sheet
        return ws
    ws = wb.create_sheet(name)
    # Remove default empty sheet if this is a brand-new workbook
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 2:
        del wb["Sheet"]
    return ws


def _write_metadata_row(ws, geo_context: GeoContext, count: int):
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    meta_text = (
        f"Source: RentCast API  |  Geo: {geo_context.label}  |  "
        f"Pulled: {timestamp}  |  Records: {count}"
    )
    ws.cell(row=1, column=1, value=meta_text)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(FIELDS))

    cell = ws.cell(row=1, column=1)
    cell.fill = FILL_META
    cell.font = FONT_META
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18


def _write_header_row(ws):
    for col_idx, field in enumerate(FIELDS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=field)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER_INPUT if field in INPUT_COLUMNS else FONT_HEADER_STD
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16


def _write_data_rows(ws, properties: list[dict]):
    for row_idx, prop in enumerate(properties, start=3):
        for col_idx, field in enumerate(FIELDS, start=1):
            value = prop.get(field)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if field == "dscr_flag":
                match value:
                    case "PASS":
                        cell.fill = FILL_PASS
                        cell.font = Font(color="FFFFFF", bold=True)
                    case "MARGINAL":
                        cell.fill = FILL_MARGINAL
                        cell.font = Font(color="000000", bold=True)
                    case "FAIL":
                        cell.fill = FILL_FAIL
                        cell.font = Font(color="FFFFFF", bold=True)


def _apply_column_widths(ws):
    for col_idx, field in enumerate(FIELDS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(field)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
